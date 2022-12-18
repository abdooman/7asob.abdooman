import openpyxl
from pathlib import Path

extend = openpyxl.load_workbook(Path.home() / Path('OneDrive', 'سطح المكتب', 'example.xlsx'))
print(extend.sheetnames)

sheetone = extend['Sheet1']
print(sheetone)


الرواتب = 0
for i in range(1,7):
    print(sheetone.cell(row=i, column=1).value, sheetone.cell(row=i, column=2).value)
    الرواتب += sheetone.cell(row=i, column=2).value

print(f'The totale of salary for employ is {الرواتب}')
     