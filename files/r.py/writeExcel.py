import pathlib
import openpyxl
from pathlib import Path 
 
#great excel file
excelgeart = openpyxl.Workbook()
print(excelgeart.sheetnames)


excelgeart.create_sheet()
excelgeart.create_sheet(title='Abdooman')
del excelgeart['Abdooman']



#tmryn
ex = excelgeart['Sheet1']
#ex['C1'] = 'Saed'
#ex['C2'] = 'Yzan'
#ex['C3'] = 'Abdullatif'
#ex['C4'] = 'Ismail'


names =['name1' , 'name2' , 'name3']
n = 0
for name in names:
    n += 1
    print(name, n)
    ex[f'C{n}'] = name

#save excel file 

excelgeart.save(filename = Path.home() / Path('OneDrive') / 'Excelfile.xlsx')


